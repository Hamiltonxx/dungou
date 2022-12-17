from config import mysqlconfig,prjcolumns
from aiohttp import web 
import aiomysql
import json
from openpyxl import load_workbook, Workbook

async def projects(request):
    conn = await aiomysql.connect(**mysqlconfig)
    cur = await conn.cursor()
    await cur.execute(f'SELECT * FROM project')
    d = await cur.fetchall()
    result = []
    for p in d:
        dt = dict(zip(prjcolumns,p))
        pics_str = dt["pics"]
        dt["pics"] = pics_str.split(";")
        result.append(dt)
    conn.close()
    return web.Response(text=json.dumps(result, ensure_ascii=False))

async def projects_search(request):
    data = await request.json()
    ground = data['ground'] if "groud" in data else 2
    shield_depth = data['shield_depth'] if "shield_depth" in data else ''
    shield_diameter = data['shield_diameter'] if "shield_diameter" in data else ''
    manufac_model = data['manufac_model'] if "manufac_model" in data else ''
    consistency_index = data['consistency_index'] if "consistency_index" in data else ''
    excavation_head = data['excavation_head'] if "excavation_head" in data else ''
    permeability = data['permeability'] if "permeability" in data else ''
    restricted_particle = data['restricted_particle'] if "restricted_particle" in data else ''
    max_particle = data['max_particle'] if "max_particle" in data else ''
    equivalent_quartz = data['equivalent_quartz'] if "equivalent_quartz" in data else ''
    province = data['province'] if "province" in data else ''

    conn = await aiomysql.connect(**mysqlconfig)
    cur = await conn.cursor()
    sql = f"SELECT * FROM project WHERE ground={ground} AND shield_depth LIKE '%{shield_depth}%' AND shield_diameter LIKE '%{shield_diameter}%' AND manufac_model LIKE '%{manufac_model}%' AND consistency_index LIKE '%{consistency_index}%' AND excavation_head LIKE '%{excavation_head}%' AND permeability LIKE '%{permeability}%' AND restricted_particle LIKE '%{restricted_particle}%' AND max_particle LIKE '%{max_particle}%' AND equivalent_quartz LIKE '%{equivalent_quartz}%' AND province LIKE '%{province}%'"
    await cur.execute(sql)
    d = await cur.fetchall()
    result = []
    for p in d:
        result.append(dict(zip(prjcolumns,p)))
    conn.close()
    return web.Response(text=json.dumps(result, ensure_ascii=False))

async def projects_export(request):
    # with open("projects_export.xlsx",'wb+') as f:
    wb = Workbook()
    sheet = wb.active
    sheet.append(("ID","工程名","施工单位","穿越地层","区间长度(m)","起止时间","盾构类型","盾构厂家/型号","土仓压力(MPa)","刀盘转速(r/min)","扭矩(kN.m)","推力(kN)","推进速度(mm/min)","土体改良","保压泵","双闸门","耐磨措施","平均进度(m/d)","刀具磨损","盾构直径(m)","盾构埋深(m)","稠度指数","开挖面水头(m)","渗透系数(cm/s)","等效石英含量(%)","限制粒径(mm)","最大粒径(mm)","省份","图片"))
    conn = await aiomysql.connect(**mysqlconfig)
    cur = await conn.cursor()
    await cur.execute("SELECT * FROM project")
    d = await cur.fetchall()
    for r in d:
        sheet.append(r)
    wb.save("projects_export.xlsx")
    conn.close()
    return web.FileResponse("projects_export.xlsx")

async def projects_import(request):
    data = await request.post()
    input_file = data['projects'].file
    content = input_file.read()
    with open("projects.xlsx", 'wb+') as f:
        f.write(content)
    wb = load_workbook("projects.xlsx")
    sheet = wb.active
    conn = await aiomysql.connect(**mysqlconfig)
    cur = await conn.cursor()
    for row in sheet.iter_rows():
        print(row[0].value, row[1].value,row[4].value)
        sql=f"UPDATE project SET shield_diameter={row[1].value},shield_depth='{row[2].value}',consistency_index={row[3].value},excavation_head={row[4].value},permeability={row[5].value},equivalent_quartz={row[6].value},restricted_particle={row[7].value},max_particle={row[8].value} WHERE id={row[0].value}"
        print(sql)
        await cur.execute(sql)
    await conn.commit()
    conn.close()
    
    return web.Response(text="OK")


async def projects_add(request):
    data = await request.json()
    name = data["name"]
    constru = data["constru"] if "constru" in data else ""
    ground = data["ground"] if "ground" in data else 2 
    length = data["length"]
    period = data["period"] if "period" in data else ""
    shield_type = data["shield_type"] if "shield_type" in data else ""
    manufac_model = data["manufac_model"] if "manufac_model" in data else ""
    soil_pressure = data["soil_pressure"] if "soil_pressure" in data else ""
    cutter_speed = data["cutter_speed"] if "cutter_speed" in data else ""
    torque = data["torque"] if "torque" in data else ""
    thrust = data["thrust"] if "thrust" in data else ""
    advance_speed = data["advance_speed"] if "advance_speed" in data else ""
    soil_improvement = data["soil_improvement"] if "soil_improvement" in data else "无"
    pressure_pump = data["pressure_pump"] if "pressure_pump" in data else "无"
    double_gate = data["double_gate"] if "double_gate" in data else "无"
    wear_resistant = data["wear_resistant"] if "wear_resistant" in data else "无"
    average_progress = data["average_progress"] if "average_progress" in data else "NULL"
    tool_wear = data["tool_wear"] if "tool_wear" in data else ""
    shield_diameter = data["shield_diameter"] if "shield_diameter" in data else "NULL"
    shield_depth = data["shield_depth"] if "shield_depth" in data else ""
    consistency_index = data["consistency_index"] if "consistency_index" in data else "NULL"
    excavation_head = data["excavation_head"] if "excavation_head" in data else "NULL"
    permeability = data["permeability"] if "permeability" in data else "NULL"
    equivalent_quartz = data["equivalent_quartz"] if "equivalent_quartz" in data else "NULL"
    restricted_particle = data["restricted_particle"] if "restricted_particle" in data else "NULL"
    max_particle = data["max_particle"] if "max_particle" in data else "NULL"
    province = data["province"] if "province" in data else ""
    pics = data["pics"] if "pics" in data else []

    pics_str = ";".join(pics)
    print(pics_str)

    conn = await aiomysql.connect(**mysqlconfig)
    cur = await conn.cursor()
    sql = f"INSERT INTO project (name,constru,ground,length,period,shield_type,manufac_model,soil_pressure,cutter_speed,torque,thrust,advance_speed,soil_improvement,pressure_pump,double_gate,wear_resistant,average_progress,tool_wear,shield_diameter,shield_depth,consistency_index,excavation_head,permeability,equivalent_quartz,restricted_particle,max_particle,province,pics) VALUES ('{name}','{constru}',{ground},{length},'{period}','{shield_type}','{manufac_model}','{soil_pressure}','{cutter_speed}','{torque}','{thrust}','{advance_speed}','{soil_improvement}','{pressure_pump}','{double_gate}','{wear_resistant}',{average_progress},'{tool_wear}',{shield_diameter},'{shield_depth}',{consistency_index},{excavation_head},{permeability},{equivalent_quartz},{restricted_particle},{max_particle},'{province}','{pics}')"
    print(sql)
    await cur.execute(sql)
    await conn.commit()
    conn.close()

    return web.Response(text="增加成功")

async def projects_delete(request):
    data = await request.json()
    ids = data["ids"]
    conn = await aiomysql.connect(**mysqlconfig)
    cur = await conn.cursor()
    if len(ids)==1:
        sql = f"DELETE FROM project WHERE id={ids[0]}"
    else:
        sql = f"DELETE FROM project WHERE id in {tuple(ids)}"
    print(sql)
    await cur.execute(sql)
    await conn.commit()
    conn.close()

    return web.Response(text="删除成功")

async def projects_update(request):
    data = await request.json()
    pics = data['pics'] if 'pics' in data else []
    pics_str = ";".join(pics)
    conn = await aiomysql.connect(**mysqlconfig)
    cur = await conn.cursor()
    sql = f'''UPDATE project set name="{data['name']}",constru="{data['constru']}",ground={data['ground']},length={data['length']},period="{data['period']}",shield_type="{data['shield_type']}",manufac_model="{data['manufac_model']}",soil_pressure="{data['soil_pressure']}",cutter_speed="{data['cutter_speed']}",torque="{data['torque']}",thrust="{data['thrust']}",advance_speed="{data['advance_speed']}",soil_improvement="{data['soil_improvement']}",pressure_pump="{data['pressure_pump']}",double_gate="{data['double_gate']}",wear_resistant="{data['wear_resistant']}",average_progress={data['average_progress']},tool_wear="{data['tool_wear']}",shield_diameter={data['shield_diameter']},shield_depth="{data['shield_depth']}",consistency_index={data['consistency_index']},excavation_head={data['excavation_head']},permeability={data['permeability']},equivalent_quartz={data['equivalent_quartz']},restricted_particle={data['restricted_particle']},max_particle={data['max_particle']},province="{data['province']}",pics="{pics_str}" WHERE id={data['id']}'''
    print(sql)
    await cur.execute(sql)
    await conn.commit()
    conn.close()

    return web.Response(text="更新成功")


